"""
Enhanced Rayne Hedge Fund Dashboard with Charts
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, AreaChart, PieChart
from datetime import datetime
from copy import copy

# Color Scheme
COLORS = {
    'bg': '0F1419',
    'accent': '4F8CFF',
    'green': '26D07C',
    'red': 'EF4F56',
    'amber': 'F5B740',
}

def add_charts_to_backtest(ws):
    """Add equity curve and metrics charts to Backtest sheet"""
    
    # Create sample equity curve data
    ws['K1'] = 'EQUITY CURVE DATA'
    ws['K1'].font = Font(bold=True, size=10, color='8B95A7')
    
    # X-axis: dates/periods
    headers = ['Period', 'Cumulative R', 'Buy & Hold']
    for col, header in enumerate(headers, 11):
        ws.cell(row=2, column=col).value = header
        ws.cell(row=2, column=col).font = Font(bold=True, color='8B95A7', size=9)
    
    # Sample equity data (50 periods)
    import random
    random.seed(42)
    cumulative_r = 0
    bh_return = 100
    
    for period in range(1, 51):
        ws.cell(row=period+2, column=11).value = period
        trade_r = random.uniform(-1, 3)
        cumulative_r += trade_r
        ws.cell(row=period+2, column=12).value = cumulative_r
        bh_return *= (1 + random.uniform(-0.01, 0.015))
        ws.cell(row=period+2, column=13).value = bh_return / 100
    
    # Equity Curve Chart
    chart1 = LineChart()
    chart1.title = "Equity Curve vs Buy & Hold"
    chart1.style = 10
    chart1.y_axis.title = 'Cumulative R'
    chart1.x_axis.title = 'Trade Period'
    chart1.height = 12
    chart1.width = 16
    
    data1 = Reference(ws, min_col=12, min_row=1, max_row=51)
    data2 = Reference(ws, min_col=13, min_row=1, max_row=51)
    chart1.add_data(data1, titles_from_data=True)
    chart1.add_data(data2, titles_from_data=True)
    
    ws.add_chart(chart1, "A26")
    
    # Win/Loss Distribution Chart
    chart2 = PieChart()
    chart2.title = "Win / Loss / Timeout Distribution"
    chart2.style = 10
    
    # Win/Loss data
    ws['K55'] = 'Outcome'
    ws['L55'] = 'Count'
    ws['K56'] = 'Wins'
    ws['L56'] = 170
    ws['K57'] = 'Losses'
    ws['L57'] = 190
    ws['K58'] = 'Timeouts'
    ws['L58'] = 52
    
    data_pie = Reference(ws, min_col=11, min_row=55, max_row=58)
    chart2.add_data(data_pie, titles_from_data=True)
    ws.add_chart(chart2, "K26")

def add_charts_to_validation(ws):
    """Add validation charts"""
    
    # Permutation test data
    ws['J2'] = 'Permutation Distribution'
    for i in range(1, 101):
        ws.cell(row=i+2, column=10).value = 0.5 + (i - 50) * 0.02  # Simulated distribution
    
    # Permutation histogram
    chart = BarChart()
    chart.title = "Permutation Test Results"
    chart.style = 10
    chart.height = 12
    chart.width = 14
    
    data = Reference(ws, min_col=10, min_row=1, max_row=101)
    chart.add_data(data, titles_from_data=True)
    
    ws.add_chart(chart, "A21")

def add_charts_to_portfolio(ws):
    """Add portfolio allocation and efficient frontier charts"""
    
    # Allocation data
    ws['I4'] = 'ES ORB'
    ws['I5'] = 'Order Block'
    ws['I6'] = 'Momentum'
    
    ws['J4'] = 40
    ws['J5'] = 30
    ws['J6'] = 30
    
    # Allocation pie chart
    chart1 = PieChart()
    chart1.title = "Portfolio Allocation by Weight"
    chart1.style = 10
    
    data = Reference(ws, min_col=9, min_row=4, max_row=6)
    vals = Reference(ws, min_col=10, min_row=4, max_row=6)
    chart1.add_data(vals)
    chart1.set_categories(data)
    chart1.height = 10
    chart1.width = 14
    
    ws.add_chart(chart1, "I9")
    
    # Risk-Return scatter (Efficient Frontier)
    ws['I15'] = 'Strategy'
    ws['J15'] = 'Risk %'
    ws['K15'] = 'Return %'
    
    strats = ['ES ORB', 'Order Block', 'Momentum', 'Portfolio']
    risks = [12.6, 15.2, 18.4, 13.8]
    returns = [22.3, 18.5, 25.1, 22.0]
    
    for idx, (s, r, ret) in enumerate(zip(strats, risks, returns), 16):
        ws.cell(row=idx, column=9).value = s
        ws.cell(row=idx, column=10).value = r
        ws.cell(row=idx, column=11).value = ret
    
    chart2 = ScatterChart()
    chart2.title = "Risk–Return Map (Efficient Frontier)"
    chart2.style = 10
    chart2.x_axis.title = "Risk (Max DD %)"
    chart2.y_axis.title = "Return (CAGR %)"
    chart2.height = 12
    chart2.width = 14
    
    xdata = Reference(ws, min_col=10, min_row=15, max_row=19)
    ydata = Reference(ws, min_col=11, min_row=15, max_row=19)
    chart2.add_data(ydata, titles_from_data=True)
    chart2.set_categories(xdata)
    
    ws.add_chart(chart2, "A21")

def enhance_workbook():
    """Load and enhance existing workbook"""
    wb = openpyxl.load_workbook('/Users/rayneperekekeme/Documents/Claude/Projects/Trading stratetgies Portfolio/Rayne Asset Management/Rayne_Hedge_Fund_Quant_Dashboard.xlsx')
    
    print("Adding charts to Backtest Lab...")
    add_charts_to_backtest(wb['Backtest Lab'])
    
    print("Adding charts to Validation Suite...")
    add_charts_to_validation(wb['Validation Suite'])
    
    print("Adding charts to Portfolio Allocator...")
    add_charts_to_portfolio(wb['Portfolio Allocator'])
    
    wb.save('/Users/rayneperekekeme/Documents/Claude/Projects/Trading stratetgies Portfolio/Rayne Asset Management/Rayne_Hedge_Fund_Quant_Dashboard.xlsx')
    print("✓ Excel dashboard enhanced with charts!")

if __name__ == '__main__':
    enhance_workbook()
